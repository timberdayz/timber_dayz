from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
import os
import shutil
import subprocess
import tempfile
from typing import Optional

import openpyxl
import pandas as pd

from modules.core.logger import get_logger

logger = get_logger(__name__)


@dataclass(frozen=True)
class SpreadsheetNormalizationResult:
    path: Path
    source_format: str
    normalized_format: str
    was_converted: bool
    converter: str
    media_stripped: bool
    cache_hit: bool = False


class SpreadsheetNormalizationService:
    def __init__(self, runtime_root: Optional[Path] = None, timeout_seconds: int = 180) -> None:
        self.runtime_root = runtime_root or (Path(tempfile.gettempdir()) / "xihong_erp_spreadsheets")
        self.runtime_root.mkdir(parents=True, exist_ok=True)
        self.timeout_seconds = timeout_seconds

    def normalize_for_runtime(
        self,
        file_path: str | Path,
        *,
        source_format: Optional[str] = None,
    ) -> SpreadsheetNormalizationResult:
        source = Path(file_path)
        if not source.exists():
            raise FileNotFoundError(source)

        normalized_source_format = str(source_format or source.suffix.lstrip(".") or "unknown").lower()
        if normalized_source_format == "xlsx":
            return SpreadsheetNormalizationResult(
                path=source,
                source_format="xlsx",
                normalized_format="xlsx",
                was_converted=False,
                converter="none",
                media_stripped=False,
                cache_hit=True,
            )

        cache_key = self._build_cache_key(source, normalized_source_format)
        output_path = self.runtime_root / f"{source.stem}.{cache_key}.normalized.xlsx"
        if output_path.exists():
            return SpreadsheetNormalizationResult(
                path=output_path,
                source_format=normalized_source_format,
                normalized_format="xlsx",
                was_converted=True,
                converter="cache",
                media_stripped=True,
                cache_hit=True,
            )

        if normalized_source_format == "html":
            self._normalize_html_table(source, output_path)
            return SpreadsheetNormalizationResult(
                path=output_path,
                source_format="html",
                normalized_format="xlsx",
                was_converted=True,
                converter="pandas_html",
                media_stripped=True,
                cache_hit=False,
            )

        if normalized_source_format not in {"xls", "xlsx_with_ole"}:
            raise RuntimeError(f"unsupported spreadsheet normalization source format: {normalized_source_format}")

        converted_path = self._convert_ole_source_to_xlsx(source, output_path.with_suffix(".converted.xlsx"))
        self._strip_workbook_to_values_only(converted_path, output_path)
        try:
            if converted_path.exists():
                converted_path.unlink()
        except OSError:
            pass

        return SpreadsheetNormalizationResult(
            path=output_path,
            source_format=normalized_source_format,
            normalized_format="xlsx",
            was_converted=True,
            converter=self._last_converter_used or "unknown",
            media_stripped=True,
            cache_hit=False,
        )

    def _build_cache_key(self, source: Path, source_format: str) -> str:
        digest = sha256()
        digest.update(source_format.encode("utf-8"))
        digest.update(str(source.resolve()).encode("utf-8"))
        stat = source.stat()
        digest.update(str(stat.st_size).encode("utf-8"))
        digest.update(str(int(stat.st_mtime)).encode("utf-8"))
        with source.open("rb") as handle:
            while True:
                chunk = handle.read(1024 * 1024)
                if not chunk:
                    break
                digest.update(chunk)
        return digest.hexdigest()[:20]

    def _normalize_html_table(self, source: Path, output_path: Path) -> None:
        dfs = pd.read_html(str(source))
        if not dfs:
            raise RuntimeError("html spreadsheet does not contain a table")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for index, df in enumerate(dfs):
                df.to_excel(writer, index=False, sheet_name=f"Sheet{index + 1}")

    def _convert_ole_source_to_xlsx(self, source: Path, converted_path: Path) -> Path:
        converted_path.parent.mkdir(parents=True, exist_ok=True)
        self._last_converter_used = None

        soffice = self._find_soffice()
        if soffice:
            try:
                result = self._convert_with_soffice(source, converted_path, soffice)
                self._last_converter_used = "soffice"
                return result
            except Exception as exc:
                logger.warning(f"Spreadsheet normalization via soffice failed: {exc}")

        if self._is_excel_com_available():
            try:
                result = self._convert_with_excel_com(source, converted_path)
                self._last_converter_used = "excel_com"
                return result
            except Exception as exc:
                logger.warning(f"Spreadsheet normalization via Excel COM failed: {exc}")

        raise RuntimeError("no spreadsheet converter available for ole xls source")

    def _find_soffice(self) -> Optional[str]:
        for candidate in (
            shutil.which("soffice"),
            shutil.which("libreoffice"),
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ):
            if candidate and Path(candidate).exists():
                return str(candidate)
        return None

    def _convert_with_soffice(self, source: Path, converted_path: Path, soffice_path: str) -> Path:
        temp_dir = converted_path.parent / f"{converted_path.stem}.soffice"
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            subprocess.run(
                [
                    soffice_path,
                    "--headless",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(temp_dir),
                    str(source),
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=self.timeout_seconds,
            )
            generated = temp_dir / f"{source.stem}.xlsx"
            if not generated.exists():
                raise RuntimeError("soffice did not produce xlsx output")
            shutil.move(str(generated), converted_path)
            return converted_path
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _is_excel_com_available(self) -> bool:
        if os.name != "nt":
            return False
        try:
            result = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    (
                        "$excel = New-Object -ComObject Excel.Application; "
                        "$v = $excel.Version; "
                        "$excel.Quit(); "
                        "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null; "
                        "Write-Output $v"
                    ),
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=30,
            )
            return bool(result.stdout.strip())
        except Exception:
            return False

    def _convert_with_excel_com(self, source: Path, converted_path: Path) -> Path:
        source_literal = str(source).replace("'", "''")
        dest_literal = str(converted_path).replace("'", "''")
        command = (
            "$ErrorActionPreference = 'Stop'; "
            "$excel = New-Object -ComObject Excel.Application; "
            "$excel.Visible = $false; "
            "$excel.DisplayAlerts = $false; "
            f"$workbook = $excel.Workbooks.Open('{source_literal}', 0, $true); "
            f"$dest = '{dest_literal}'; "
            "$workbook.SaveAs($dest, 51); "
            "$workbook.Close($false); "
            "$excel.Quit(); "
            "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null; "
            "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null"
        )
        subprocess.run(
            ["powershell", "-NoProfile", "-Command", command],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=self.timeout_seconds,
        )
        if not converted_path.exists():
            raise RuntimeError("excel com did not produce xlsx output")
        return converted_path

    def _strip_workbook_to_values_only(self, source_xlsx: Path, output_path: Path) -> None:
        workbook = openpyxl.load_workbook(source_xlsx, data_only=True, read_only=False)
        try:
            stripped = openpyxl.Workbook()
            default_sheet = stripped.active
            stripped.remove(default_sheet)

            for worksheet in workbook.worksheets:
                target = stripped.create_sheet(title=worksheet.title[:31] or "Sheet")
                for row in worksheet.iter_rows(values_only=True):
                    target.append(list(row))
            stripped.save(output_path)
            stripped.close()
        finally:
            workbook.close()


_service: SpreadsheetNormalizationService | None = None


def get_spreadsheet_normalization_service() -> SpreadsheetNormalizationService:
    global _service
    if _service is None:
        _service = SpreadsheetNormalizationService()
    return _service
