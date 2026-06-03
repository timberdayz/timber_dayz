from pathlib import Path

import openpyxl
import pytest

from backend.services.spreadsheet_normalization_service import SpreadsheetNormalizationService


def test_normalize_html_xls_to_runtime_xlsx(tmp_path):
    source = tmp_path / "sample.xls"
    source.write_text(
        """
        <html><body>
        <table>
          <tr><th>SKU</th><th>库存</th></tr>
          <tr><td>A-1</td><td>8</td></tr>
        </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    service = SpreadsheetNormalizationService(runtime_root=tmp_path / "normalized")
    result = service.normalize_for_runtime(source, source_format="html")

    assert result.was_converted is True
    assert result.normalized_format == "xlsx"
    assert result.media_stripped is True
    assert result.path.exists()

    workbook = openpyxl.load_workbook(result.path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet["A1"].value == "SKU"
    assert sheet["A2"].value == "A-1"
    assert str(sheet["B2"].value) == "8"
    workbook.close()


def test_normalize_ole_xls_uses_converter_then_strips_to_values(tmp_path, monkeypatch):
    source = tmp_path / "sample.xls"
    source.write_bytes(b"\xD0\xCF\x11\xE0demo")

    converted = tmp_path / "converted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "库存"
    sheet["A1"] = "SKU"
    sheet["B1"] = "库存"
    sheet["A2"] = "A-1"
    sheet["B2"] = 10
    workbook.save(converted)
    workbook.close()

    service = SpreadsheetNormalizationService(runtime_root=tmp_path / "normalized")

    monkeypatch.setattr(service, "_convert_with_soffice", lambda *_args, **_kwargs: converted)
    monkeypatch.setattr(service, "_find_soffice", lambda: "soffice")

    result = service.normalize_for_runtime(source, source_format="xls")

    assert result.was_converted is True
    assert result.converter == "soffice"
    assert result.media_stripped is True
    assert result.path.exists()
    assert result.path.suffix.lower() == ".xlsx"
    assert result.path != converted

    stripped = openpyxl.load_workbook(result.path, data_only=True)
    sheet = stripped[stripped.sheetnames[0]]
    assert sheet["A2"].value == "A-1"
    assert sheet["B2"].value == 10
    stripped.close()


def test_normalize_for_runtime_reuses_cached_result(tmp_path, monkeypatch):
    source = tmp_path / "sample.xls"
    source.write_bytes(b"\xD0\xCF\x11\xE0demo")

    converted = tmp_path / "converted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "SKU"
    workbook.save(converted)
    workbook.close()

    service = SpreadsheetNormalizationService(runtime_root=tmp_path / "normalized")

    calls = {"convert": 0}

    def _fake_convert(*_args, **_kwargs):
        calls["convert"] += 1
        return converted

    monkeypatch.setattr(service, "_convert_with_soffice", _fake_convert)
    monkeypatch.setattr(service, "_find_soffice", lambda: "soffice")

    first = service.normalize_for_runtime(source, source_format="xls")
    second = service.normalize_for_runtime(source, source_format="xls")

    assert calls["convert"] == 1
    assert first.path == second.path
    assert second.cache_hit is True


def test_normalize_for_runtime_raises_when_no_converter_available(tmp_path, monkeypatch):
    source = tmp_path / "sample.xls"
    source.write_bytes(b"\xD0\xCF\x11\xE0demo")

    service = SpreadsheetNormalizationService(runtime_root=tmp_path / "normalized")
    monkeypatch.setattr(service, "_find_soffice", lambda: None)
    monkeypatch.setattr(service, "_is_excel_com_available", lambda: False)

    with pytest.raises(RuntimeError, match="converter"):
        service.normalize_for_runtime(source, source_format="xls")
